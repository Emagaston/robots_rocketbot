import xlwings as xw
import time, smtplib, os
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from openpyxl import load_workbook
from email import policy


email = os.environ.get('EmailT')
email_password = os.environ.get('PassT')

archivo_excel= "D:\Documentos\Python_xlwings\Base Seguimiento Observ Auditoría al_30042021.xlsx"

open_excel= xw.Book(archivo_excel)

libro_trabajo = load_workbook(archivo_excel, data_only=True)

hoja = libro_trabajo.active

url= "https://roc.myrb.io/s1/forms/M6I8P2PDOZFDBYYG"

# Configuración de Selenium
driver = webdriver.Chrome()  

time.sleep(5)

columna = hoja['J']

for cell in columna[1:]:  #excluye cabecera

    estado_proceso = cell.value  
    fila= cell.row

    dato_columna_pro = hoja.cell(row=fila, column=1).value
    dato_columna_obs = hoja.cell(row=fila, column=2).value
    dato_columna_t_riesgo= hoja.cell(row=fila, column=3).value
    dato_columna_sev = hoja.cell(row=fila, column=4).value
    dato_columna_fec = hoja.cell(row=fila, column=6).value
    dato_columna_resp = hoja.cell(row=fila, column=7).value

    

    if estado_proceso == 'Regularizado':
        # Abre el formulario web
        driver.get(url)

        campo_dato_t_riesgo = driver.find_element(By.ID, 'tipo_riesgo')
        campo_dato_t_riesgo.send_keys(str(dato_columna_t_riesgo))

        campo_dato_resp = driver.find_element(By.ID, 'res')
        campo_dato_resp.send_keys(str(dato_columna_resp))

        campo_dato_obs = driver.find_element(By.ID, 'obs')
        campo_dato_obs.send_keys(str(dato_columna_obs))

        campo_dato_fec = driver.find_element(By.ID, 'date')
        campo_dato_fec.send_keys(dato_columna_fec.strftime("%d/%m/%Y"))

        campo_dato_pro = Select(driver.find_element(By.ID, 'process'))
        for option in campo_dato_pro.options:
            if dato_columna_pro in option.text:
                option.click()
                break 

        campo_dato_sev= Select(driver.find_element(By.ID, 'severidad'))
        for option in campo_dato_sev.options:
            if dato_columna_sev in option.text:
                option.click()
                break 

        # Envia el formulario
        submit_button = driver.find_element(By.ID, 'submit')
        submit_button.click()

    elif estado_proceso == 'Atrasado':
        dato_columna_e_resp = hoja.cell(row=fila, column=9).value
        dato_sin_espacios = dato_columna_e_resp.replace(" ","")
        
        # Prepara el correo electrónico
        msg = MIMEMultipart(policy=policy.default)
        msg['From'] = email
        msg['To'] = dato_sin_espacios
        msg['Subject'] = 'Proceso Atrasado'
        
        # Cuerpo del correo electrónico - (Procesos, Estado, Observacion, fecha compromiso)
        body = f"""\
        El proceso {dato_columna_pro} está atrasado.
        Estado: {estado_proceso}
        Observación: {dato_columna_obs}
        Fecha de Compromiso: {dato_columna_fec.strftime("%d/%m/%Y")}
        """
        msg.attach(MIMEText(body, 'plain'))
        
        # Enviar el correo electrónico
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(email, email_password)
        server.send_message(msg)
        server.quit()

driver.quit()